"""피드백 벡터 클러스터링 → 지라 태스크 생성

1. 피드백 항목에 대해 LLM이 세부 증상 자유 서술
2. 서술을 Bedrock Titan으로 임베딩
3. Agglomerative Clustering으로 유사도 기반 그룹핑
4. 클러스터별 지라 태스크 구조로 엑셀 출력
"""
import sys
import os
import json
import time
import logging
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed

import boto3
import numpy as np
from sklearn.cluster import AgglomerativeClustering
from sklearn.metrics.pairwise import cosine_distances

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s")
logger = logging.getLogger(__name__)

# ── 설정 ──
DATA_PATH = "data/classified_data/v5_2026-03-23.json"
OUTPUT_DIR = "./exports"
OUTPUT_FILE = "tech_issues_clustered_2026-03-23.xlsx"
DISTANCE_THRESHOLD = 0.60  # 코사인 거리 임계값 (낮을수록 엄격)

# ── Step 1: LLM 세부 증상 서술 ──
DESCRIBE_SYSTEM_PROMPT = """당신은 금융 교육 플랫폼의 VOC 분석가입니다.

아래 사용자 VOC 텍스트를 읽고, **기술/서비스 이슈를 한 문장으로 구체적으로 서술**하세요.

규칙:
- 증상 중심으로 서술 (원인 추측 X)
- 어떤 기능/화면에서, 어떤 증상이 발생했는지
- 정보가 부족해도 최선의 추측으로 서술. "구체적 증상 불명"이라고 쓰지 말 것
  - "안되요" → "앱 특정 기능 작동 불가"
  - "접속이 안돼요" → "앱/웹 접속 불가"
  - "라이브특강 안되는 건가요" → "라이브 특강 접속 불가"
- 기술 이슈든 운영 이슈든 서비스 피드백이든 모두 서술
- 환불/해지/구독/결제 → "구독 해지 경로 불명확", "환불 요청 후 처리 지연"
- 운영 정책 → "운영진 대응 지연", "커뮤니티 관리 부재"
- 기능 요청 → "댓글 기능 부재", "검색 필터 개선 필요"
- 한 문장, 50자 이내

예시:
- "앱 푸시 알림 설정 ON인데 새 글 알림 미수신"
- "줌 라이브 강의 입장 버튼 클릭 시 무반응"
- "구독 해지 버튼 접근 경로 불명확, 해지 방법 안내 부재"
- "환불 요청 후 7일 경과에도 처리 미완료"
- "커뮤니티 댓글 기능 없어 사용자 간 소통 불가"

응답: 한 문장만 출력 (JSON 아님)"""


def describe_issues(items):
    """각 피드백 항목에 대해 LLM이 세부 증상 서술"""
    bedrock = boto3.client("bedrock-runtime", region_name="us-west-2")
    model_id = "us.anthropic.claude-haiku-4-5-20251001-v1:0"

    def describe_single(text):
        for attempt in range(3):
            try:
                resp = bedrock.invoke_model(
                    modelId=model_id,
                    body=json.dumps({
                        "anthropic_version": "bedrock-2023-05-31",
                        "max_tokens": 100,
                        "system": DESCRIBE_SYSTEM_PROMPT,
                        "messages": [{"role": "user", "content": text[:500]}],
                    }),
                )
                result = json.loads(resp["body"].read())
                return result["content"][0]["text"].strip()
            except Exception as e:
                if "Throttl" in str(e) and attempt < 2:
                    time.sleep(3 * (attempt + 1))
                    continue
                logger.warning(f"서술 실패: {e}")
                return f"서술 실패: {text[:50]}"

    logger.info(f"세부 증상 서술 시작: {len(items)}건")
    start = time.time()

    with ThreadPoolExecutor(max_workers=10) as executor:
        future_map = {}
        for i, item in enumerate(items):
            future = executor.submit(describe_single, item['text'])
            future_map[future] = i

        for future in as_completed(future_map):
            idx = future_map[future]
            items[idx]['issue_description'] = future.result()

    logger.info(f"서술 완료: {time.time()-start:.1f}초")
    return items


# ── Step 2: 벡터 임베딩 ──
def embed_descriptions(items):
    """Bedrock Titan으로 임베딩"""
    bedrock = boto3.client("bedrock-runtime", region_name="us-west-2")

    def embed_single(text):
        for attempt in range(3):
            try:
                resp = bedrock.invoke_model(
                    modelId="amazon.titan-embed-text-v2:0",
                    body=json.dumps({
                        "inputText": text[:2000],
                        "dimensions": 256,
                    }),
                )
                result = json.loads(resp["body"].read())
                return result["embedding"]
            except Exception as e:
                if "Throttl" in str(e) and attempt < 2:
                    time.sleep(3 * (attempt + 1))
                    continue
                logger.warning(f"임베딩 실패: {e}")
                return None

    logger.info(f"임베딩 시작: {len(items)}건")
    start = time.time()

    embeddings = []
    valid_items = []

    with ThreadPoolExecutor(max_workers=10) as executor:
        future_map = {}
        for i, item in enumerate(items):
            # 서술 + 원문을 concat → 서술이 달라도 원문 유사도로 보완
            desc = item.get('issue_description', '')
            text = item.get('text', '')[:200]
            embed_text = f"{desc} | {text}"
            future = executor.submit(embed_single, embed_text)
            future_map[future] = i

        for future in as_completed(future_map):
            idx = future_map[future]
            emb = future.result()
            if emb is not None:
                embeddings.append(emb)
                valid_items.append(items[idx])

    logger.info(f"임베딩 완료: {len(embeddings)}건, {time.time()-start:.1f}초")
    return valid_items, np.array(embeddings)


# ── Step 3: 클러스터링 ──
def _keyword_overlap_matrix(items):
    """아이템 간 키워드(tags) 겹침 비율 행렬 계산"""
    n = len(items)
    overlap = np.zeros((n, n))
    tag_sets = []
    for item in items:
        tags = item.get('tags', [])
        if isinstance(tags, str):
            tags = [t.strip() for t in tags.split(',')]
        # tags를 개별 단어로 분해 (복합 태그 "수업접속" → "수업", "접속")
        words = set()
        for tag in tags:
            # 공백/·으로 분리
            for part in tag.replace('·', ' ').replace('_', ' ').split():
                if len(part) >= 2:
                    words.add(part)
        tag_sets.append(words)

    for i in range(n):
        for j in range(i + 1, n):
            if not tag_sets[i] or not tag_sets[j]:
                continue
            intersection = len(tag_sets[i] & tag_sets[j])
            union = len(tag_sets[i] | tag_sets[j])
            if union > 0:
                jaccard = intersection / union
                overlap[i][j] = jaccard
                overlap[j][i] = jaccard

    return overlap


def cluster_issues(items, embeddings, distance_threshold=DISTANCE_THRESHOLD):
    """Agglomerative Clustering — 임베딩 유사도 + 키워드 겹침 보정"""
    cosine_dist = cosine_distances(embeddings)
    keyword_overlap = _keyword_overlap_matrix(items)

    # 키워드 겹침이 있으면 거리를 줄임
    # 최종 거리 = 코사인 거리 × (1 - keyword_weight × 겹침률)
    keyword_weight = 0.4
    dist_matrix = cosine_dist * (1 - keyword_weight * keyword_overlap)

    clustering = AgglomerativeClustering(
        n_clusters=None,
        distance_threshold=distance_threshold,
        metric='precomputed',
        linkage='average',
    )
    labels = clustering.fit_predict(dist_matrix)

    n_clusters = len(set(labels))
    logger.info(f"클러스터링 완료: {n_clusters}개 클러스터 (threshold={distance_threshold})")

    # 클러스터별 그룹핑
    clusters = defaultdict(list)
    for item, label in zip(items, labels):
        item['cluster_id'] = int(label)
        clusters[int(label)].append(item)

    return clusters


# ── Step 4: 지라 태스크 생성 + 엑셀 ──
def generate_jira_excel(clusters, output_path, short_items=None):
    """클러스터별 지라 태스크 구조로 엑셀 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)

    # ── Sheet 1: Jira Tasks (클러스터 = 태스크) ──
    ws1 = wb.active
    ws1.title = "Jira Tasks"

    headers = ['클러스터', '건수', '대표 증상', '영향 커뮤니티', '감성 분포',
               '날짜 범위', 'subtag 분포', '대표 원문 1', '대표 원문 2',
               'Jira 제목 제안', 'Jira 설명']

    for col, name in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # 클러스터를 건수 내림차순 정렬
    sorted_clusters = sorted(clusters.items(), key=lambda x: -len(x[1]))

    row = 2
    for cluster_id, items in sorted_clusters:
        # 대표 증상 (가장 많이 나오는 description 패턴)
        descriptions = [i.get('issue_description', '') for i in items]
        # 가장 긴 (= 가장 구체적인) description을 대표로
        rep_desc = max(descriptions, key=len) if descriptions else ''

        # 영향 커뮤니티
        masters = sorted(set(i.get('masterName', '') for i in items))
        masters_str = ', '.join(masters[:5])
        if len(masters) > 5:
            masters_str += f' 외 {len(masters)-5}개'

        # 감성 분포
        sentiments = Counter(i.get('sentiment', '중립') for i in items)
        sent_str = ', '.join(f"{s}:{c}" for s, c in sentiments.most_common())

        # 날짜 범위
        dates = sorted(i.get('createdAt', '') for i in items if i.get('createdAt'))
        date_range = f"{dates[0]} ~ {dates[-1]}" if dates else ''

        # subtag 분포
        subtags = Counter(i.get('subtag', '') for i in items)
        subtag_str = ', '.join(f"{s}:{c}" for s, c in subtags.most_common())

        # 대표 원문 (가장 구체적인 것)
        sorted_by_len = sorted(items, key=lambda x: len(x.get('text', '')), reverse=True)
        quote1 = sorted_by_len[0]['text'][:250] if len(sorted_by_len) > 0 else ''
        quote2 = sorted_by_len[1]['text'][:250] if len(sorted_by_len) > 1 else ''

        # 부정 비율로 우선순위 추정
        neg_ratio = sentiments.get('부정', 0) / len(items)
        if len(items) >= 5 and neg_ratio > 0.5:
            priority = "P1"
        elif len(items) >= 3:
            priority = "P2"
        else:
            priority = "P3"

        # Jira 제목
        jira_title = f"[{priority}] {rep_desc} ({len(items)}건)"

        # Jira 설명
        jira_desc_lines = [
            f"## 증상",
            f"{rep_desc}",
            f"",
            f"## 영향 범위",
            f"- 보고 건수: {len(items)}건",
            f"- 영향 커뮤니티: {masters_str}",
            f"- 발생 기간: {date_range}",
            f"- 감성: {sent_str}",
            f"",
            f"## 사용자 원문 (상위 3건)",
        ]
        for q in sorted_by_len[:3]:
            jira_desc_lines.append(f"- \"{q['text'][:150]}\" ({q.get('masterName','')}, {q.get('createdAt','')})")

        jira_desc = '\n'.join(jira_desc_lines)

        values = [
            f"C{cluster_id}", len(items), rep_desc, masters_str, sent_str,
            date_range, subtag_str, quote1, quote2, jira_title, jira_desc,
        ]

        for col, val in enumerate(values, 1):
            cell = ws1.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=(col >= 8))

        row += 1

    col_widths = [10, 8, 40, 25, 20, 22, 20, 45, 45, 45, 60]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row-1}"
    ws1.freeze_panes = 'A2'

    # ── Sheet 2: 전체 상세 ──
    ws2 = wb.create_sheet("전체 상세")

    headers2 = ['클러스터', '증상 서술', 'subtag', '출처', '마스터',
                 '감성', '원문', '작성일']
    for col, name in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    all_items = []
    for cluster_id, items in sorted_clusters:
        for item in items:
            all_items.append(item)

    for row_idx, item in enumerate(all_items, 2):
        values = [
            f"C{item.get('cluster_id', '?')}", item.get('issue_description', ''),
            item.get('subtag', ''), item.get('type', ''),
            item.get('masterName', ''), item.get('sentiment', ''),
            item.get('text', '')[:300], item.get('createdAt', ''),
        ]
        for col, val in enumerate(values, 1):
            cell = ws2.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=(col in [2, 7]))

    col_widths_2 = [10, 40, 12, 8, 10, 8, 50, 12]
    for i, w in enumerate(col_widths_2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}{len(all_items)+1}"
    ws2.freeze_panes = 'A2'

    # ── Sheet 3: 통계 ──
    ws3 = wb.create_sheet("통계")
    ws3.cell(row=1, column=1, value="항목").font = Font(bold=True)
    ws3.cell(row=1, column=2, value="값").font = Font(bold=True)

    stats = [
        ("전체 피드백", len(all_items)),
        ("클러스터 수", len(clusters)),
        ("평균 클러스터 크기", f"{len(all_items)/len(clusters):.1f}건"),
        ("최대 클러스터", f"{max(len(v) for v in clusters.values())}건"),
        ("1건 클러스터 (개별건)", sum(1 for v in clusters.values() if len(v) == 1)),
    ]
    for r, (label, val) in enumerate(stats, 2):
        ws3.cell(row=r, column=1, value=label)
        ws3.cell(row=r, column=2, value=val)
    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 15

    # ── Sheet 4: 미분류 (짧은 원문) ──
    if short_items:
        ws4 = wb.create_sheet("미분류")
        headers4 = ['subtag', '출처', '마스터', '감성', '원문', '작성일']
        for col, name in enumerate(headers4, 1):
            cell = ws4.cell(row=1, column=col, value=name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        for row_idx, item in enumerate(short_items, 2):
            values = [
                item.get('subtag', ''), item.get('type', ''),
                item.get('masterName', ''), item.get('sentiment', ''),
                item.get('text', ''), item.get('createdAt', ''),
            ]
            for col, val in enumerate(values, 1):
                cell = ws4.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border

        ws4.column_dimensions['A'].width = 12
        ws4.column_dimensions['E'].width = 30
        ws4.column_dimensions['F'].width = 12

    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    wb.save(output_path)
    logger.info(f"엑셀 저장: {output_path}")

    return len(clusters), len(all_items)


def main():
    print("=" * 60)
    print("피드백 벡터 클러스터링 → 지라 태스크")
    print("=" * 60)

    # 데이터 로드
    with open(DATA_PATH) as f:
        data = json.load(f)

    feedback = [i for i in data['items'] if i['topic'] == '피드백']
    # 이용문의 제외 — 단순 문의는 클러스터링 대상 아님
    feedback = [i for i in feedback if i.get('subtag') != '이용문의']
    logger.info(f"피드백 {len(feedback)}건 로드 (이용문의 제외)")

    # Step 1: LLM 세부 증상 서술
    feedback = describe_issues(feedback)

    # 짧은 원문(10자 이하) 분리 — 분류 불가
    short_items = [i for i in feedback if len(i.get('text', '').strip()) <= 10]
    tech_items = [i for i in feedback if len(i.get('text', '').strip()) > 10]
    logger.info(f"클러스터링 대상: {len(tech_items)}건, 짧은 원문 제외: {len(short_items)}건")

    # Step 2: 임베딩
    tech_items, embeddings = embed_descriptions(tech_items)

    # Step 3: 클러스터링
    clusters = cluster_issues(tech_items, embeddings)

    # Step 4: 엑셀 생성
    output_path = os.path.join(OUTPUT_DIR, OUTPUT_FILE)
    n_clusters, n_items = generate_jira_excel(clusters, output_path, short_items=short_items)

    print(f"\n{'='*60}")
    print(f"완료!")
    print(f"  피드백: {len(feedback)}건")
    print(f"  기술이슈: {n_items}건")
    print(f"  클러스터 (= 지라 태스크): {n_clusters}개")
    print(f"  엑셀: {output_path}")
    print(f"{'='*60}")

    # 슬랙 전송
    try:
        from src.integrations.slack_client import SlackNotifier
        slack = SlackNotifier()
        msg = slack._send_message("[3/23 주차] 기술 이슈 클러스터링 리포트")
        if msg.get("ok"):
            thread_ts = msg["ts"]

            # 클러스터 요약
            sorted_clusters = sorted(clusters.items(), key=lambda x: -len(x[1]))
            summary = ["피드백 벡터 클러스터링 결과:\n"]
            for cid, items in sorted_clusters[:10]:
                desc = items[0].get('issue_description', '')[:60]
                summary.append(f"  C{cid}: {len(items)}건 — {desc}")
            summary.append(f"\n총 {n_clusters}개 클러스터, {n_items}건")

            slack._send_message('\n'.join(summary), thread_ts=thread_ts)
            slack.upload_file_to_thread(
                output_path, thread_ts,
                title="기술이슈 클러스터링 리포트",
                comment="벡터 유사도 기반 클러스터링 → Jira 태스크 매핑"
            )
            print("슬랙 전송 완료!")
    except Exception as e:
        logger.error(f"슬랙 전송 실패: {e}")


if __name__ == "__main__":
    main()
