"""1~3월 기술 이슈 세분류 리포트 생성

1. BigQuery에서 1~3월 편지/게시글 전체 조회
2. v5 Bedrock Haiku로 분류
3. 피드백 항목 추출 → 기술이슈 세분류 (2차)
4. 지라 매핑 구조로 엑셀 생성
5. 슬랙으로 전송
"""
import sys
import os
import json
import time
import logging
from datetime import datetime, timedelta
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed

import boto3

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from src.bigquery.client import BigQueryClient
from src.bigquery.queries import WeeklyDataQuery
from src.classifier_v5.bedrock_classifier import BedrockV5Classifier
from src.integrations.slack_client import SlackNotifier

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s")
logger = logging.getLogger(__name__)

# ── 설정 ──
START_DATE = "2026-01-01"
END_DATE = "2026-02-01"  # exclusive — 1월만 먼저
OUTPUT_DIR = "./exports"
OUTPUT_FILE = "tech_issues_2026_01.xlsx"

# 기술이슈 세분류 프롬프트
TECH_ISSUE_SYSTEM_PROMPT = """당신은 금융 교육 플랫폼의 기술 이슈 분류기입니다.

아래 VOC 텍스트를 읽고, **기술 이슈 유형**을 판단하세요.

## 이슈 유형 (정확히 하나 선택)
- 수업접속불가: 줌/라이브/강의 접속 안됨, 입장 불가
- 앱크래시: 화면 안뜸, 흰화면, 앱 멈춤, 앱 재설치로 해결
- 알림미작동: 푸시알림 안옴, 알림설정 했는데 안됨
- 링크오류: 링크 클릭 불가, 링크 안열림
- 글수정오류: 게시글/댓글 수정 안됨
- 결제프로세스오류: 결제 시 오류, 주소 입력 불가, 결제 실패
- 인증오류: 인증문자 안옴, 회원가입 안됨, 로그인 불가
- 미디어오류: 오디오 안됨, 영상 안나옴, 음질 불량
- 서버장애: 사이트 전체 접속 불가, 서버 다운
- 콘텐츠접근불가: 녹화본/강의 삭제됨, 콘텐츠 안보임, 수강 내역 미표시
- 버튼UI미작동: 특정 버튼 안 눌림, UI 요소 반응 없음, 제출/신청 불가
- 속도저하: 앱/웹 느림, 로딩 오래 걸림, 응답 지연
- 해지프로세스오류: 구독 해지 안됨, 해지 버튼 안 보임, 해지 후 과금
- 기타기술: 위에 해당 안되는 기술 문제 → reason 필드에 구체적 증상 필수 기재

## 우선순위 판단
- P0_긴급: 다수 사용자 영향, 서비스 중단 수준
- P1_높음: 핵심 기능 장애 (결제, 수업 접속)
- P2_보통: 불편하지만 우회 가능
- P3_낮음: 사소한 UI/기능 이슈

## 응답: JSON만 출력
{"issue_type": "수업접속불가", "priority": "P1_높음", "affected_feature": "줌 라이브", "detail": "줌 접속 불가 - 스마트폰/PC 모두"}"""

FEATURE_REQUEST_SYSTEM_PROMPT = """당신은 금융 교육 플랫폼의 기능요청 분류기입니다.

아래 VOC 텍스트를 읽고, **기능 요청 유형**을 판단하세요.

## 요청 유형 (정확히 하나 선택)
- 댓글기능: 댓글/답글 기능 요청
- 검색개선: 콘텐츠 검색, 기업 검색 개선
- 카테고리정리: 콘텐츠 분류/카테고리 추가
- 자동재생: 오디오/영상 자동재생
- VOD요청: 녹화본/VOD 업로드 요청
- UI개선: 화면 레이아웃, 사용성 개선
- 알림개선: 알림 설정, 공지 기능
- 기타요청: 위에 해당 안되는 기능 요청 → reason 필드에 구체적 내용 필수 기재

## 우선순위 판단
- P1_높음: 다수 사용자가 반복 요청
- P2_보통: 합리적 개선 요청
- P3_낮음: 개인적 선호

## 응답: JSON만 출력
{"request_type": "댓글기능", "priority": "P1_높음", "detail": "커뮤니티 댓글 기능 부재로 소통 불가"}"""


# ── Step 1: BigQuery 조회 ──
def fetch_all_data(start_date, end_date):
    """1~3월 편지/게시글 전체 조회 + 마스터 정보 매핑"""
    client = BigQueryClient()
    query = WeeklyDataQuery(client)

    logger.info(f"BigQuery 조회: {start_date} ~ {end_date}")

    # 마스터 정보
    master_info = query.get_master_info()
    logger.info(f"마스터 {len(master_info)}명 로드")

    # 편지글
    letters = query.get_weekly_letters(start_date, end_date)
    for item in letters:
        mid = item.get('masterId')
        if mid and mid in master_info:
            item['masterName'] = master_info[mid]['displayName']
            item['masterClubName'] = master_info[mid]['clubName']
        else:
            item['masterName'] = 'Unknown'
            item['masterClubName'] = 'Unknown'

    # 게시글
    posts = query.get_weekly_posts(start_date, end_date)

    # postBoardId → masterId 매핑
    board_query = f"""
    SELECT _id as boardId, masterId
    FROM `{client.project_id}.us_plus_new.postboards`
    """
    board_to_master = {b['boardId']: b['masterId']
                       for b in client.execute_query(board_query)}

    for item in posts:
        board_id = item.get('postBoardId')
        actual_master_id = board_to_master.get(board_id, board_id)
        if actual_master_id and actual_master_id in master_info:
            item['masterName'] = master_info[actual_master_id]['displayName']
            item['masterClubName'] = master_info[actual_master_id]['clubName']
        else:
            item['masterName'] = 'Unknown'
            item['masterClubName'] = 'Unknown'

    logger.info(f"조회 완료: 편지 {len(letters)}건, 게시글 {len(posts)}건")
    return letters, posts


# ── Step 2: v5 분류 ──
def classify_all(letters, posts):
    """v5 Bedrock Haiku로 전체 분류"""
    classifier = BedrockV5Classifier(max_workers=10)

    all_items = []

    # 편지글 분류
    if letters:
        logger.info(f"편지글 {len(letters)}건 분류 시작...")
        for item in letters:
            text = item.get('message', '')
            result = classifier.classify_single(text)
            all_items.append({
                '_id': item.get('_id', ''),
                'type': 'letter',
                'masterName': item.get('masterName', ''),
                'masterClubName': item.get('masterClubName', ''),
                'text': text,
                'createdAt': str(item.get('createdAt', ''))[:10],
                **result,
            })
        logger.info(f"편지글 분류 완료")

    # 게시글 분류
    if posts:
        logger.info(f"게시글 {len(posts)}건 분류 시작...")
        for item in posts:
            text = item.get('textBody', '') or item.get('body', '')
            result = classifier.classify_single(text)
            all_items.append({
                '_id': item.get('_id', ''),
                'type': 'post',
                'masterName': item.get('masterName', ''),
                'masterClubName': item.get('masterClubName', ''),
                'text': text,
                'createdAt': str(item.get('createdAt', ''))[:10],
                **result,
            })
        logger.info(f"게시글 분류 완료")

    cost = classifier.get_cost_report()
    logger.info(f"1차 분류 비용: ${cost['cost_usd']}, 에러: {cost['errors']}")

    return all_items, cost


def classify_all_batch(letters, posts):
    """v5 Bedrock Haiku 병렬 분류 (ThreadPoolExecutor)"""
    classifier = BedrockV5Classifier(max_workers=10)

    # 통합 리스트 구성
    raw_items = []
    for item in letters:
        raw_items.append({
            '_id': item.get('_id', ''),
            'type': 'letter',
            'masterName': item.get('masterName', ''),
            'masterClubName': item.get('masterClubName', ''),
            'text': item.get('message', ''),
            'createdAt': str(item.get('createdAt', ''))[:10],
        })
    for item in posts:
        raw_items.append({
            '_id': item.get('_id', ''),
            'type': 'post',
            'masterName': item.get('masterName', ''),
            'masterClubName': item.get('masterClubName', ''),
            'text': item.get('textBody', '') or item.get('body', ''),
            'createdAt': str(item.get('createdAt', ''))[:10],
        })

    logger.info(f"전체 {len(raw_items)}건 병렬 분류 시작 (workers=10)...")
    start_time = time.time()
    done = 0

    with ThreadPoolExecutor(max_workers=10) as executor:
        future_map = {}
        for i, item in enumerate(raw_items):
            future = executor.submit(classifier.classify_single, item['text'])
            future_map[future] = i

        for future in as_completed(future_map):
            idx = future_map[future]
            result = future.result()
            raw_items[idx].update(result)
            done += 1
            if done % 500 == 0 or done == len(raw_items):
                elapsed = time.time() - start_time
                rate = done / elapsed if elapsed > 0 else 0
                logger.info(f"  {done}/{len(raw_items)} ({rate:.1f}건/초)")

    cost = classifier.get_cost_report()
    logger.info(f"1차 분류 완료: ${cost['cost_usd']}, 에러: {cost['errors']}")

    return raw_items, cost


# ── Step 3: 피드백 기술이슈 세분류 ──
def classify_tech_issues(feedback_items):
    """피드백 항목 중 기술이슈 세분류"""
    bedrock = boto3.client("bedrock-runtime", region_name="us-west-2")
    model_id = "us.anthropic.claude-haiku-4-5-20251001-v1:0"

    # 순수 기술이슈만 — 결제·구독/환불·해지는 CS 이슈이므로 제외
    tech_subtags = {'접속·오류', '기능요청', 'UX개선'}

    tech_items = [i for i in feedback_items if i.get('subtag') in tech_subtags]
    logger.info(f"기술이슈 세분류 대상: {len(tech_items)}건")

    def classify_single(item):
        text = item.get('text', '')[:500]
        subtag = item.get('subtag', '')

        if subtag in {'접속·오류', 'UX개선'}:
            prompt = TECH_ISSUE_SYSTEM_PROMPT
            parse_key = 'issue_type'
        elif subtag == '기능요청':
            prompt = FEATURE_REQUEST_SYSTEM_PROMPT
            parse_key = 'request_type'
        else:
            # 결제·구독, 환불·해지는 기술 이슈인지 아닌지 판단
            prompt = TECH_ISSUE_SYSTEM_PROMPT
            parse_key = 'issue_type'

        for attempt in range(3):
            try:
                resp = bedrock.invoke_model(
                    modelId=model_id,
                    body=json.dumps({
                        "anthropic_version": "bedrock-2023-05-31",
                        "max_tokens": 200,
                        "system": prompt,
                        "messages": [{"role": "user", "content": text}],
                    }),
                )
                result = json.loads(resp["body"].read())
                raw = result["content"][0]["text"].strip()

                # JSON 파싱
                start = raw.find("{")
                end = raw.rfind("}") + 1
                if start >= 0 and end > start:
                    parsed = json.loads(raw[start:end])
                    return parsed
                return {}
            except Exception as e:
                if "Throttl" in str(e) and attempt < 2:
                    time.sleep(5 * (attempt + 1))
                    continue
                logger.warning(f"세분류 실패: {e}")
                return {}

    start_time = time.time()
    done = 0

    with ThreadPoolExecutor(max_workers=10) as executor:
        future_map = {}
        for i, item in enumerate(tech_items):
            future = executor.submit(classify_single, item)
            future_map[future] = i

        for future in as_completed(future_map):
            idx = future_map[future]
            result = future.result()
            tech_items[idx]['tech_detail'] = result
            done += 1
            if done % 100 == 0 or done == len(tech_items):
                logger.info(f"  세분류 {done}/{len(tech_items)}")

    elapsed = time.time() - start_time
    logger.info(f"세분류 완료: {elapsed:.1f}초")

    return tech_items


# ── Step 4: 엑셀 생성 ──
def generate_excel(all_items, tech_items, output_path):
    """지라 매핑 구조 엑셀 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)

    priority_fills = {
        'P0_긴급': PatternFill(start_color='FF4444', end_color='FF4444', fill_type='solid'),
        'P1_높음': PatternFill(start_color='FF8800', end_color='FF8800', fill_type='solid'),
        'P2_보통': PatternFill(start_color='FFCC00', end_color='FFCC00', fill_type='solid'),
        'P3_낮음': PatternFill(start_color='88CC88', end_color='88CC88', fill_type='solid'),
    }
    priority_fonts = {
        'P0_긴급': Font(color='FFFFFF', bold=True),
        'P1_높음': Font(color='FFFFFF', bold=True),
        'P2_보통': Font(bold=True),
        'P3_낮음': Font(),
    }

    # ── Sheet 1: 지라 이슈 요약 (클러스터링) ──
    ws1 = wb.active
    ws1.title = "이슈 요약 (Jira Epic)"

    # 이슈 클러스터링
    clusters = defaultdict(list)
    for item in tech_items:
        detail = item.get('tech_detail', {})
        issue_type = detail.get('issue_type') or detail.get('request_type') or '미분류'
        clusters[issue_type].append(item)

    headers = ['우선순위', '이슈 유형', '건수', '영향 커뮤니티', '기간 집중도',
               '핵심 증상', '대표 원문 1', '대표 원문 2', 'Jira Epic 제안']
    for col, name in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    row = 2
    # 우선순위 순 정렬
    def cluster_priority(items):
        priorities = [i.get('tech_detail', {}).get('priority', 'P3_낮음') for i in items]
        p_order = {'P0_긴급': 0, 'P1_높음': 1, 'P2_보통': 2, 'P3_낮음': 3}
        min_p = min(p_order.get(p, 3) for p in priorities)
        return min_p, -len(items)

    for issue_type, items in sorted(clusters.items(), key=lambda x: cluster_priority(x[1])):
        # 우선순위: 가장 높은 것
        priorities = [i.get('tech_detail', {}).get('priority', 'P3_낮음') for i in items]
        p_order = {'P0_긴급': 0, 'P1_높음': 1, 'P2_보통': 2, 'P3_낮음': 3}
        top_priority = min(priorities, key=lambda p: p_order.get(p, 3))

        # 영향 커뮤니티
        masters = set(i.get('masterName', '') for i in items if i.get('masterName'))
        masters_str = ', '.join(sorted(masters)[:5])
        if len(masters) > 5:
            masters_str += f' 외 {len(masters)-5}개'

        # 기간 집중도 (월별)
        months = Counter()
        for i in items:
            date_str = i.get('createdAt', '')[:7]  # YYYY-MM
            if date_str:
                months[date_str] += 1
        period_str = ', '.join(f"{m}: {c}건" for m, c in sorted(months.items()))

        # 핵심 증상
        details = [i.get('tech_detail', {}).get('detail', '') for i in items if i.get('tech_detail', {}).get('detail')]
        symptom = details[0] if details else ''

        # 대표 원문 (가장 구체적인 것 = 텍스트 길이 기준)
        sorted_by_len = sorted(items, key=lambda x: len(x.get('text', '')), reverse=True)
        quote1 = sorted_by_len[0]['text'][:200] if len(sorted_by_len) > 0 else ''
        quote2 = sorted_by_len[1]['text'][:200] if len(sorted_by_len) > 1 else ''

        # Jira epic 제안
        affected = '플랫폼 전체' if len(masters) >= 3 else masters_str
        epic_title = f"[{top_priority}] {issue_type} ({len(items)}건, {affected})"

        values = [top_priority, issue_type, len(items), masters_str, period_str,
                  symptom, quote1, quote2, epic_title]

        for col, val in enumerate(values, 1):
            cell = ws1.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if col == 1:
                fill = priority_fills.get(top_priority)
                font = priority_fonts.get(top_priority)
                if fill:
                    cell.fill = fill
                if font:
                    cell.font = font

        row += 1

    col_widths_1 = [12, 18, 8, 25, 25, 35, 45, 45, 40]
    for i, w in enumerate(col_widths_1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row-1}"
    ws1.freeze_panes = 'A2'

    # ── Sheet 2: 전체 기술이슈 상세 (Jira Story) ──
    ws2 = wb.create_sheet("기술이슈 상세 (Jira Story)")

    headers2 = ['우선순위', '이슈유형', 'subtag', '출처', '마스터', '클럽',
                 '감성', '세부내용', '원문', '작성일', 'Jira Story 제안']
    for col, name in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 우선순위 → 이슈유형 → 날짜 순 정렬
    p_order = {'P0_긴급': 0, 'P1_높음': 1, 'P2_보통': 2, 'P3_낮음': 3}
    sorted_items = sorted(tech_items, key=lambda x: (
        p_order.get(x.get('tech_detail', {}).get('priority', 'P3_낮음'), 3),
        x.get('tech_detail', {}).get('issue_type') or x.get('tech_detail', {}).get('request_type') or '',
        x.get('createdAt', ''),
    ))

    for row_idx, item in enumerate(sorted_items, 2):
        detail = item.get('tech_detail', {})
        priority = detail.get('priority', 'P3_낮음')
        issue_type = detail.get('issue_type') or detail.get('request_type') or '미분류'
        detail_text = detail.get('detail', '')

        # Jira story 제안
        story_title = f"{issue_type}: {detail_text[:60]}"

        values = [
            priority, issue_type, item.get('subtag', ''),
            item.get('type', ''), item.get('masterName', ''),
            item.get('masterClubName', ''), item.get('sentiment', ''),
            detail_text, item.get('text', '')[:300],
            item.get('createdAt', ''), story_title,
        ]

        for col, val in enumerate(values, 1):
            cell = ws2.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=(col in [8, 9, 11]))
            if col == 1:
                fill = priority_fills.get(priority)
                font = priority_fonts.get(priority)
                if fill:
                    cell.fill = fill
                if font:
                    cell.font = font

    col_widths_2 = [12, 16, 12, 8, 10, 18, 8, 35, 50, 12, 40]
    for i, w in enumerate(col_widths_2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}{len(sorted_items)+1}"
    ws2.freeze_panes = 'A2'

    # ── Sheet 3: 피드백 전체 (비기술 포함) ──
    ws3 = wb.create_sheet("피드백 전체")

    feedback_items = [i for i in all_items if i.get('topic') == '피드백']

    headers3 = ['subtag', '출처', '마스터', '클럽', '감성', '요약', '태그', '원문', '작성일']
    for col, name in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    feedback_items.sort(key=lambda x: (x.get('subtag', ''), x.get('createdAt', '')))

    for row_idx, item in enumerate(feedback_items, 2):
        values = [
            item.get('subtag', ''), item.get('type', ''),
            item.get('masterName', ''), item.get('masterClubName', ''),
            item.get('sentiment', ''), item.get('summary', ''),
            ', '.join(item.get('tags', [])),
            item.get('text', '')[:300], item.get('createdAt', ''),
        ]
        for col, val in enumerate(values, 1):
            cell = ws3.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=(col == 8))

    col_widths_3 = [12, 8, 10, 18, 8, 40, 25, 50, 12]
    for i, w in enumerate(col_widths_3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(headers3))}{len(feedback_items)+1}"
    ws3.freeze_panes = 'A2'

    # ── Sheet 4: 통계 요약 ──
    ws4 = wb.create_sheet("통계 요약")

    ws4.cell(row=1, column=1, value="구분").font = Font(bold=True)
    ws4.cell(row=1, column=2, value="값").font = Font(bold=True)

    total = len(all_items)
    feedback_count = len(feedback_items)
    tech_count = len(tech_items)

    stats = [
        ("조회 기간", f"{START_DATE} ~ {END_DATE}"),
        ("전체 건수", total),
        ("피드백 건수", f"{feedback_count} ({feedback_count/total*100:.1f}%)"),
        ("기술이슈 건수", f"{tech_count} ({tech_count/total*100:.1f}%)"),
        ("", ""),
        ("--- 이슈유형별 ---", ""),
    ]

    # 이슈유형별 통계
    type_counts = Counter()
    for item in tech_items:
        detail = item.get('tech_detail', {})
        t = detail.get('issue_type') or detail.get('request_type') or '미분류'
        type_counts[t] += 1
    for t, c in type_counts.most_common():
        stats.append((t, f"{c}건"))

    stats.append(("", ""))
    stats.append(("--- 월별 피드백 추이 ---", ""))
    month_counts = Counter()
    for item in feedback_items:
        m = item.get('createdAt', '')[:7]
        if m:
            month_counts[m] += 1
    for m, c in sorted(month_counts.items()):
        stats.append((m, f"{c}건"))

    for row_idx, (label, val) in enumerate(stats, 2):
        ws4.cell(row=row_idx, column=1, value=label)
        ws4.cell(row=row_idx, column=2, value=val)

    ws4.column_dimensions['A'].width = 25
    ws4.column_dimensions['B'].width = 20

    # 저장
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    logger.info(f"엑셀 저장: {output_path}")

    return feedback_count, tech_count


# ── Step 5: 슬랙 전송 ──
def send_to_slack(output_path, feedback_count, tech_count, clusters_summary):
    """슬랙에 메시지 + 엑셀 파일 전송"""
    slack = SlackNotifier()

    # 메인 메시지
    main_msg = f"[Q1 2026] 기술 이슈 세분류 리포트"

    main_response = slack._send_message(main_msg)
    if not main_response.get("ok"):
        logger.error(f"슬랙 메시지 실패: {main_response}")
        return

    thread_ts = main_response.get("ts")

    # 요약 스레드
    summary_lines = [
        f"1~3월 편지/게시글 기술 이슈 세분류 결과입니다.",
        f"",
        f"피드백 전체: {feedback_count}건",
        f"기술이슈 (개발 대응 필요): {tech_count}건",
        f"",
        f"--- 이슈 유형별 Top 5 ---",
    ]
    for issue_type, count in clusters_summary[:5]:
        summary_lines.append(f"  {issue_type}: {count}건")

    summary_lines.append(f"\n상세 내용은 첨부 엑셀을 확인해주세요.")

    slack._send_message('\n'.join(summary_lines), thread_ts=thread_ts)

    # 파일 업로드
    result = slack.upload_file_to_thread(
        file_path=output_path,
        thread_ts=thread_ts,
        title="Q1 2026 기술이슈 세분류 리포트",
        comment="Jira Epic/Story 매핑 포함"
    )
    if result.get("ok"):
        logger.info(f"슬랙 파일 업로드 완료: {result.get('file_url')}")
    else:
        logger.error(f"파일 업로드 실패: {result}")


def main():
    print("=" * 60)
    print("Q1 2026 기술 이슈 세분류 리포트 생성")
    print("=" * 60)

    output_path = os.path.join(OUTPUT_DIR, OUTPUT_FILE)
    cache_path = os.path.join(OUTPUT_DIR, "q1_2026_classified_all.json")

    # 캐시 확인
    if os.path.exists(cache_path):
        logger.info(f"캐시 로드: {cache_path}")
        with open(cache_path) as f:
            all_items = json.load(f)
        logger.info(f"캐시에서 {len(all_items)}건 로드")
    else:
        # Step 1: BigQuery 조회
        letters, posts = fetch_all_data(START_DATE, END_DATE)

        # Step 2: v5 분류
        all_items, cost = classify_all_batch(letters, posts)

        # 캐시 저장
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        with open(cache_path, 'w') as f:
            json.dump(all_items, f, ensure_ascii=False, indent=2)
        logger.info(f"캐시 저장: {cache_path} ({len(all_items)}건)")

    # Step 3: 피드백 기술이슈 세분류
    feedback_items = [i for i in all_items if i.get('topic') == '피드백']
    logger.info(f"피드백 {len(feedback_items)}건 중 기술이슈 세분류...")

    tech_items = classify_tech_issues(feedback_items)

    # Step 4: 엑셀 생성
    feedback_count, tech_count = generate_excel(all_items, tech_items, output_path)

    # 클러스터 요약
    type_counts = Counter()
    for item in tech_items:
        detail = item.get('tech_detail', {})
        t = detail.get('issue_type') or detail.get('request_type') or '미분류'
        type_counts[t] += 1
    clusters_summary = type_counts.most_common()

    print(f"\n{'='*60}")
    print(f"완료!")
    print(f"  전체: {len(all_items)}건")
    print(f"  피드백: {feedback_count}건")
    print(f"  기술이슈: {tech_count}건")
    print(f"  엑셀: {output_path}")
    print(f"{'='*60}")

    # Step 5: 슬랙 전송
    try:
        send_to_slack(output_path, feedback_count, tech_count, clusters_summary)
        print("슬랙 전송 완료!")
    except Exception as e:
        logger.error(f"슬랙 전송 실패: {e}")
        print(f"슬랙 전송 실패 (엑셀은 생성됨): {e}")


if __name__ == "__main__":
    main()
